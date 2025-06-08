import os
from flask import Flask, send_file
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from datetime import datetime

app = Flask(__name__)

# Nastavení připojení k MySQL
db_config = {
    'host': 'sql4.webzdarma.cz',
    'user': 'collatzwzcz4682',
    'password': 'V63AK*&G5,0b(3,eyd.2',
    'database': 'collatzwzcz4682'
}

@app.route('/export', methods=['GET'])
def export_xlsx():
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT player_id, player_name FROM players")
    players = cursor.fetchall()
    id_to_name = {p['player_id']: p['player_name'] for p in players}
    player_names = list(id_to_name.values())

    # Create matrix
    size = len(player_names)
    matrix = [[''] + player_names]
    name_to_index = {name: idx for idx, name in enumerate(player_names)}

    # Initialize rows
    for name in player_names:
        row = [name] + ['' for _ in range(size)]
        matrix.append(row)

    # Load match data
    cursor.execute("SELECT * FROM matches WHERE match_status='pending'")
    matches = cursor.fetchall()
    for match in matches:
        p1 = id_to_name[match['player1_id']]
        p2 = id_to_name[match['player2_id']]
        i = name_to_index[p1] + 1
        j = name_to_index[p2] + 1
        matrix[i][j] = '✖️'

    cursor.close()
    conn.close()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Turnaj"

    # Nadpis
    title = f"Turnaj {datetime.now().year}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=size + 1)
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Tabulka
    for row_idx, row_data in enumerate(matrix, start=3):
        for col_idx, cell in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell)

    # Stylování
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Výstup do paměti
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output,
                     as_attachment=True,
                     download_name=f"turnaj_{datetime.now().year}.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))  # Render nastaví PORT automaticky
    app.run(host="0.0.0.0", port=port)
